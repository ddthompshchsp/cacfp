
import io
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
import pytz

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as XLImage

# ------------------------------
# App config
# ------------------------------
st.set_page_config(page_title="CACFP Full Roster (Signatures)", layout="wide")
st.title("CACFP Enrollment Roster — Full Extract (with Signatures)")
st.caption("Builds one row per child. If signatures can't be embedded, writes '[signature image]' instead. Missing fields = red ✗. Enrollment Date after Parent/Guardian Date = red.")

uploaded = st.file_uploader("Upload CACFP Enrollment workbook (.xlsx)", type=["xlsx"])

# ------------------------------
# Schema / constants
# ------------------------------
TARGET_COLS = [
    "PID",
    "First Name",
    "Last Name",
    "Center",
    "Enrollment Date",
    "Parent Signature",
    "Parent/Guardian Name",
    "Parent/Guardian Date",
    "Staff Signature",
    "Staff Name",
    "Staff Date",
]

DATE_PAT = r"(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})"
RED = "C00000"

# ------------------------------
# Helpers
# ------------------------------
def _clean_date(s: Any) -> str:
    if s is None or (isinstance(s, float) and pd.isna(s)) or (isinstance(s, str) and s.strip() == ""):
        return ""
    try:
        dt = pd.to_datetime(str(s), errors="coerce")
        if pd.isna(dt):
            m = re.search(DATE_PAT, str(s))
            if not m:
                return ""
            dt = pd.to_datetime(m.group(1), errors="coerce")
        return dt.strftime("%m/%d/%Y")
    except Exception:
        m = re.search(DATE_PAT, str(s))
        return m.group(1) if m else ""

def _text_at_right(ws: Worksheet, r: int, c: int, max_shift: int = 6) -> Optional[str]:
    for cc in range(c + 1, min(ws.max_column, c + max_shift) + 1):
        v = ws.cell(row=r, column=cc).value
        if isinstance(v, str) and v.strip():
            return v.strip()
        if isinstance(v, (int, float)):
            return str(v)
    return None

def _find_label(ws: Worksheet, label_substr: str, limit_rows: int = 120) -> Optional[Tuple[int, int]]:
    lab = label_substr.lower()
    for r in range(1, min(ws.max_row, limit_rows) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and lab in v.lower():
                return r, c
    return None

def _split_name(full: str) -> Tuple[str, str]:
    if not full or not str(full).strip():
        return "", ""
    parts = str(full).strip().split()
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]

def _extract_center(ws: Worksheet) -> str:
    # Look in the top banner for a string starting with HCHSP, then take the next segment as Center
    for r in range(1, min(ws.max_row, 20) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() and v.strip().lower().startswith("hchsp"):
                s = v.strip()
                parts = re.split(r"[:\-–|]", s, maxsplit=3)
                if len(parts) >= 2:
                    center = parts[1].strip()
                    if center:
                        return center
    return ""

def _anchor_to_rc(anchor) -> Tuple[int, int]:
    try:
        if isinstance(anchor, str):
            m = re.match(r"([A-Za-z]+)(\d+)", anchor)
            if m:
                col = column_index_from_string(m.group(1))
                row = int(m.group(2))
                return row, col
        if hasattr(anchor, "row") and hasattr(anchor, "col_idx"):
            return int(anchor.row), int(anchor.col_idx)
        if hasattr(anchor, "_from"):
            fm = anchor._from
            return int(fm.row + 1), int(fm.col + 1)
    except Exception:
        pass
    return 1, 1

def _nearest_label(ws: Worksheet, row: int, col: int, phrases: List[str], scan_rows: int = 12) -> Optional[str]:
    # Upwards
    for rr in range(max(1, row - scan_rows), row + 1):
        v = ws.cell(row=rr, column=col).value
        if isinstance(v, str):
            s = v.strip().lower()
            for p in phrases:
                if p in s:
                    return p
    # Leftwards
    for cc in range(max(1, col - 12), col + 1):
        v = ws.cell(row=row, column=cc).value
        if isinstance(v, str):
            s = v.strip().lower()
            for p in phrases:
                if p in s:
                    return p
    return None

def _image_to_bytes(img) -> Optional[bytes]:
    """Best-effort: pull embedded image bytes from an openpyxl Image object."""
    # Some openpyxl builds expose ._data(); if not, we can't extract reliably here.
    try:
        raw = img._data()
        if raw:
            return raw
    except Exception:
        pass
    return None

def _extract_form(ws: Worksheet) -> Dict[str, Any]:
    """Extract one child's record from a form-like sheet."""
    row: Dict[str, Any] = {k: "" for k in TARGET_COLS}
    row["Parent Signature"] = None
    row["Staff Signature"] = None

    # PID
    pid = ""
    for r in range(1, min(ws.max_row, 120) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "pid" in v.lower():
                right = _text_at_right(ws, r, c, max_shift=4)
                if right and re.search(r"\d{5,10}", right.replace(" ", "")):
                    pid = re.search(r"\d{5,10}", right.replace(" ", "")).group(0)
                    break
        if pid:
            break
    row["PID"] = pid

    # Child Name -> split into First/Last
    child = ""
    pos = _find_label(ws, "name")
    if pos:
        rr, cc = pos
        val = _text_at_right(ws, rr, cc, max_shift=6)
        if val:
            child = val
    first, last = _split_name(child)
    row["First Name"], row["Last Name"] = first, last

    # Center
    row["Center"] = _extract_center(ws)

    # Enrollment Date
    enroll = ""
    pos_e = _find_label(ws, "enrollment date")
    if pos_e:
        rr, cc = pos_e
        v = ws.cell(row=rr, column=cc + 1).value
        enroll = _clean_date(v if v else ws.cell(row=rr, column=cc).value)
    row["Enrollment Date"] = enroll

    # Parent/Guardian name + date
    pg_name = ""
    pg_date = ""
    pos_p = _find_label(ws, "parent/guardian")
    if pos_p:
        rr, cc = pos_p
        val = _text_at_right(ws, rr, cc, max_shift=6)
        if val:
            pg_name = val
        vd = ws.cell(row=rr, column=cc + 1).value or ws.cell(row=rr, column=cc + 2).value
        pg_date = _clean_date(vd)
    row["Parent/Guardian Name"] = pg_name
    row["Parent/Guardian Date"] = pg_date

    # Staff name + date
    staff_name = ""
    staff_date = ""
    pos_s = _find_label(ws, "staff:")
    if pos_s:
        rr, cc = pos_s
        val = _text_at_right(ws, rr, cc, max_shift=6)
        if val:
            staff_name = val
        vd = ws.cell(row=rr, column=cc + 1).value or ws.cell(row=rr, column=cc + 2).value
        staff_date = _clean_date(vd)
    row["Staff Name"] = staff_name
    row["Staff Date"] = staff_date

    # Signatures (images) near labels
    parent_img = None
    staff_img = None
    for img in getattr(ws, "_images", []):
        rr, cc = _anchor_to_rc(img.anchor)
        lbl = _nearest_label(ws, rr, cc, ["parent signature", "staff signature"], scan_rows=12)
        if lbl == "parent signature" and parent_img is None:
            parent_img = img
        elif lbl == "staff signature" and staff_img is None:
            staff_img = img
    row["Parent Signature"] = parent_img
    row["Staff Signature"] = staff_img

    return row

def _export_roster(df: pd.DataFrame) -> bytes:
    """Export roster DataFrame to Excel with images (when possible) and validations/styling."""
    wb = Workbook()
    ws = wb.active
    ws.title = "CACFP Roster"

    header_row = 3
    data_row0 = header_row + 1

    tz = pytz.timezone("America/Chicago")
    now_str = datetime.now(tz).strftime("%m/%d/%Y %I:%M %p CT")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(TARGET_COLS))
    ws["A1"] = f"CACFP Enrollment Roster — Exported {now_str}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for j, col in enumerate(TARGET_COLS, start=1):
        ws.cell(row=header_row, column=j, value=col).alignment = Alignment(horizontal="center", vertical="center")

    col_idx = {name: TARGET_COLS.index(name) + 1 for name in TARGET_COLS}

    # Write rows
    for i, rec in df.iterrows():
        r = data_row0 + i

        # text fields
        for name in ["PID","First Name","Last Name","Center","Enrollment Date",
                     "Parent/Guardian Name","Parent/Guardian Date","Staff Name","Staff Date"]:
            val = str(rec.get(name) or "")
            cell = ws.cell(row=r, column=col_idx[name], value=val if val else "✗")
            if not val:
                cell.font = Font(bold=True, color=RED)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

        # signatures
        for name in ["Parent Signature", "Staff Signature"]:
            c = col_idx[name]
            obj = rec.get(name)
            if obj is None:
                # If we found the *label* but no image object, write placeholder text
                # Heuristic: if Parent/Guardian or Staff name/date exists, assume signature likely exists
                related_has_text = False
                if name.startswith("Parent"):
                    related_has_text = bool(str(rec.get("Parent/Guardian Name") or "").strip())
                else:
                    related_has_text = bool(str(rec.get("Staff Name") or "").strip())
                if related_has_text:
                    ws.cell(row=r, column=c, value="[signature image]").alignment = Alignment(vertical="center")
                else:
                    cell = ws.cell(row=r, column=c, value="✗")
                    cell.font = Font(bold=True, color=RED)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                raw = _image_to_bytes(obj)
                if raw:
                    stream = io.BytesIO(raw)
                    new_img = XLImage(stream)
                    new_img.width = 120
                    new_img.height = 28
                    ws.add_image(new_img, f"{get_column_letter(c)}{r}")
                else:
                    ws.cell(row=r, column=c, value="[signature image]").alignment = Alignment(vertical="center")

        # validation: Enrollment Date AFTER Parent/Guardian Date -> make Enrollment Date red
        try:
            ed = pd.to_datetime(rec.get("Enrollment Date"), errors="coerce")
            pd1 = pd.to_datetime(rec.get("Parent/Guardian Date"), errors="coerce")
            if pd.notna(ed) and pd.notna(pd1) and ed > pd1:
                ws.cell(row=r, column=col_idx["Enrollment Date"]).font = Font(bold=True, color=RED)
        except Exception:
            pass

    # Excel Table
    max_row = ws.max_row
    max_col = len(TARGET_COLS)
    ref = f"A{header_row}:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName="CACFPRoster", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    ws.freeze_panes = f"A{data_row0}"

    # autosize columns
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = len(str(ws.cell(row=header_row, column=col).value) or "")
        for r in range(data_row0, max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 3, 48)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

# ------------------------------
# Main
# ------------------------------
if uploaded:
    tmp = Path(str(Path.cwd() / f"__cacfp_{datetime.now().timestamp()}.xlsx"))
    tmp.write_bytes(uploaded.read())

    wb = load_workbook(tmp, data_only=True)
    rows: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        rec = _extract_form(ws)
        # keep if PID or a name exists
        if str(rec.get("PID","")).strip() or str(rec.get("First Name","")).strip() or str(rec.get("Last Name","")).strip():
            rows.append(rec)

    df = pd.DataFrame(rows, columns=TARGET_COLS)

    # normalize dates
    for col in ["Enrollment Date","Parent/Guardian Date","Staff Date"]:
        if col in df.columns:
            df[col] = df[col].map(_clean_date)

    # ensure all columns exist
    for col in TARGET_COLS:
        if col not in df.columns:
            df[col] = "" if "Signature" not in col else None
    df = df[TARGET_COLS]

    if df.empty:
        st.error("No child records found. Make sure each child is on its own sheet with an HCHSP header and signature sections.")
    else:
        st.success(f"Extracted {len(df):,} child record(s). Review and download below.")
        st.dataframe(df.drop(columns=["Parent Signature","Staff Signature"]).head(20), use_container_width=True)
        xlsx = _export_roster(df)
        st.download_button(
            "⬇️ Download CACFP Roster (.xlsx)",
            data=xlsx,
            file_name=f"CACFP_Roster_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.info("Upload the CACFP Enrollment workbook (.xlsx) to begin.")
