import io
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
import pytz

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as XLImage

st.set_page_config(page_title="CACFP Roster — Initial Report Extractor", layout="wide")

# ---------------- Header (logo+title) ----------------
logo_path = Path("header_logo.png")
hdr_l, hdr_c, hdr_r = st.columns([1, 2, 1])
with hdr_c:
    if logo_path.exists():
        st.image(str(logo_path), width=320)
    st.markdown(
        """
        <h1 style='text-align:center; margin: 8px 0 4px;'>CACFP Roster — Initial Report Extractor</h1>
        <p style='text-align:center; font-size:16px; margin-top:0;'>
        Upload the initial (tabular) report. We'll auto-detect the header row and map the columns.<br>
        Missing fields → <b style='color:#C00000'>red ✗</b>. Enrollment Date after Parent/Guardian Date → <b style='color:#C00000'>red</b>.
        </p>
        """,
        unsafe_allow_html=True,
    )
st.divider()

TARGET_COLS = [
    "PID",
    "First Name",
    "Last Name",
    "Center",
    "Enrollment Date",
    "Parent Signature",
    "Parent/Guardian Name",
    "Parent/Guardian Date",
    "Staff Signature",
    "Staff Name",
    "Staff Date",
]

DATE_PAT = r"(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})"
RED = "C00000"

def _clean_date(s):
    if s is None or (isinstance(s, float) and pd.isna(s)) or (isinstance(s, str) and s.strip() == ""):
        return ""
    try:
        dt = pd.to_datetime(str(s), errors="coerce")
        if pd.isna(dt):
            m = re.search(DATE_PAT, str(s))
            if not m:
                return ""
            dt = pd.to_datetime(m.group(1), errors="coerce")
        return dt.strftime("%m/%d/%Y")
    except Exception:
        m = re.search(DATE_PAT, str(s))
        return m.group(1) if m else ""

def _split_name(full):
    if not full or not str(full).strip():
        return "", ""
    parts = str(full).strip().split()
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]

def _rename_columns(cols):
    mapping = {}
    for c in cols:
        s = str(c).strip()
        sl = s.lower()
        if re.search(r"\bparticipant\b.*\bpid\b|\bpid\b", sl):
            mapping[c] = "PID"
        elif re.search(r"\bchild\b.*\bname\b|\bname of child\b|\bchildname\b", sl):
            mapping[c] = "Child Name"
        elif re.search(r"\bfirst\b.*\bname\b", sl):
            mapping[c] = "First Name"
        elif re.search(r"\blast\b.*\bname\b", sl):
            mapping[c] = "Last Name"
        elif re.search(r"\bcenter\b|\bschool\b", sl):
            mapping[c] = "Center"
        elif re.search(r"\benrollment\b.*\bdate\b|\bentry\b.*\bdate\b", sl):
            mapping[c] = "Enrollment Date"
        elif re.search(r"\bparent\b.*\bguardian\b.*\bname\b", sl):
            mapping[c] = "Parent/Guardian Name"
        elif re.search(r"\bparent\b.*\bguardian\b.*\bdate\b", sl):
            mapping[c] = "Parent/Guardian Date"
        elif re.search(r"\bstaff\b.*\bname\b", sl):
            mapping[c] = "Staff Name"
        elif re.search(r"\bstaff\b.*\bdate\b", sl):
            mapping[c] = "Staff Date"
        elif re.search(r"\bparent\b.*signature|\bparent/guardian\b.*signature", sl):
            mapping[c] = "Parent Signature"
        elif re.search(r"\bstaff\b.*signature", sl):
            mapping[c] = "Staff Signature"
        else:
            mapping[c] = s
    return [mapping[c] for c in cols]

def _detect_header_row(df):
    for r in range(min(60, len(df))):
        row = df.iloc[r, :].astype(str).str.strip().str.lower().tolist()
        hits = sum(
            1 for x in row if re.search(
                r"pid|participant|child.*name|first.*name|last.*name|center|school|enrollment.*date|parent/guardian|staff|date|signature", x
            )
        )
        if hits >= 3:
            return r
    return 0

def _export(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "CACFP Roster"

    header_row = 3
    data_row0 = header_row + 1

    tz = pytz.timezone("America/Chicago")
    now_str = datetime.now(tz).strftime("%m/%d/%Y %I:%M %p CT")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(TARGET_COLS))
    ws["A1"] = f"CACFP Enrollment Roster — Exported {now_str}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for j, col in enumerate(TARGET_COLS, start=1):
        ws.cell(row=header_row, column=j, value=col).alignment = Alignment(horizontal="center", vertical="center")

    col_idx = {name: TARGET_COLS.index(name) + 1 for name in TARGET_COLS}

    for i, rec in df.iterrows():
        r = data_row0 + i
        for name in TARGET_COLS:
            val = str(rec.get(name) or "")
            cell = ws.cell(row=r, column=col_idx[name], value=val if val else "✗")
            if not val:
                cell.font = Font(bold=True, color=RED)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")
        # Validation
        try:
            ed = pd.to_datetime(rec.get("Enrollment Date"), errors="coerce")
            pd1 = pd.to_datetime(rec.get("Parent/Guardian Date"), errors="coerce")
            if pd.notna(ed) and pd.notna(pd1) and ed > pd1:
                ws.cell(row=r, column=col_idx["Enrollment Date"]).font = Font(bold=True, color=RED)
        except Exception:
            pass

    # Table + banding + filters
    ref = f"A{header_row}:{get_column_letter(len(TARGET_COLS))}{ws.max_row}"
    table = Table(displayName="CACFPRoster", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    ws.freeze_panes = f"A{data_row0}"

    # autosize
    for col in range(1, len(TARGET_COLS) + 1):
        letter = get_column_letter(col)
        max_len = len(str(ws.cell(row=header_row, column=col).value) or "")
        for r in range(data_row0, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 3, 48)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

# ---------------- UI: Upload + parsing options ----------------
up = st.file_uploader("Upload the initial (tabular) report (.xlsx)", type=["xlsx"])

if up:
    # Read sheet names
    xls = pd.ExcelFile(up)
    sheet = st.selectbox("Pick the sheet that has the roster table:", xls.sheet_names)

    raw = pd.read_excel(up, sheet_name=sheet, header=None, dtype=str)
    guess = _detect_header_row(raw)
    hdr = st.number_input("Header row (0-indexed)", min_value=0, max_value=max(0, len(raw)-1), value=int(guess), step=1,
                          help="This is the row index where the column labels are.")

    # Parse using the chosen header row
    df = pd.read_excel(up, sheet_name=sheet, header=int(hdr), dtype=str)
    df = df.dropna(axis=1, how="all").dropna(how="all")
    df.columns = _rename_columns(df.columns)

    # If only Child Name provided, split into First/Last
    if "Child Name" in df.columns and ("First Name" not in df.columns or "Last Name" not in df.columns):
        fn, ln = zip(*[ _split_name(v) for v in df["Child Name"].tolist() ])
        if "First Name" not in df.columns:
            df["First Name"] = list(fn)
        if "Last Name" not in df.columns:
            df["Last Name"] = list(ln)

    # Normalize common date fields
    for col in ["Enrollment Date", "Parent/Guardian Date", "Staff Date"]:
        if col in df.columns:
            df[col] = df[col].map(_clean_date)

    # Build final frame
    out = pd.DataFrame()
    for col in TARGET_COLS:
        out[col] = df[col] if col in df.columns else ""

    st.subheader("Preview")
    st.dataframe(out[["PID","First Name","Last Name","Center","Enrollment Date","Parent/Guardian Name","Parent/Guardian Date","Staff Name","Staff Date"]].head(20),
                 use_container_width=True)

    xlsx = _export(out)
    st.success(f"Parsed {len(out):,} rows from '{sheet}'.")
    st.download_button(
        "⬇️ Download CACFP Roster (.xlsx)",
        data=xlsx,
        file_name=f"CACFP_Roster_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

else:
    st.info("Upload the initial report (.xlsx) to begin.")
